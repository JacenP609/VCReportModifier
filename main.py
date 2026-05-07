import os
import re
import json
import time
import shutil
import random
import logging
import datetime
from pathlib import Path
from datetime import timedelta
from typing import Dict, List, Optional, Tuple, Set

from bs4 import BeautifulSoup

import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter


# ================================================================
# CONFIG
# ================================================================

ORIGINAL_ROOT = r"C:\Users\hyper.park\PycharmProjects\VCReportModifier\Security"
NEW_ROOT = r"C:\Users\hyper.park\PycharmProjects\VCReportModifier\SecurityFixed"

SWE4_JSON_ROOT = r"C:\Users\hyper.park\PycharmProjects\VCReportModifier\SWE4_Json"

# Excel extraction input.
# If you run HTML modification first, keep this as NEW_ROOT.
INPUT_ROOT = NEW_ROOT

OUTPUT_XLSX_PATH = r"C:\Users\hyper.park\PycharmProjects\VCReportModifier\Security_TCNames.xlsx"

OVERWRITE_NEW_ROOT = True
RECURSIVE_SEARCH = True

LOG_PATH = "UnitTC.log"
ENABLE_TIME_LOG = True

TEST_START_DATE = "01 MAY 2026  1:00:00 PM"
EXECUTION_START_DATE = "04 MAY 2026  9:00:00 AM"

DATE_FORMAT = "%d %b %Y  %I:%M:%S %p"

RANDOM_SEED = 20260506

CREATION_MIN_STEP_MINUTES = 10
CREATION_MAX_STEP_MINUTES = 20

EXECUTION_STEP_PER_TC_COUNT = 5
EXECUTION_STEP_MINUTES = 1

JSON_FILE_MAP = {
    "hil": "HIL_FunctionList.json",
    "fil": "FIL_FunctionList.json",
    "security": "Security_FunctionList.json",
    "ftl": "FTL_FunctionList.json",
}

JSON_CACHE: Dict[str, Dict[str, Dict[str, List[str]]]] = {}


# ================================================================
# LOGGING / EXCEL STYLE
# ================================================================

logging.basicConfig(
    filename=LOG_PATH,
    filemode="w",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s %(message)s"
)

HEADER_FILL = PatternFill("solid", start_color="38E3FF")
HEADER_FONT = Font(bold=True, size=12, color="0000FF")
THIN_SIDE = Side(border_style="thin", color="000000")
HEADER_BORDER = Border(
    top=THIN_SIDE,
    left=THIN_SIDE,
    right=THIN_SIDE,
    bottom=THIN_SIDE
)
BODY_BORDER = Border(
    top=THIN_SIDE,
    left=THIN_SIDE,
    right=THIN_SIDE,
    bottom=THIN_SIDE
)


# ================================================================
# COMMON TEXT UTIL
# ================================================================

def clean_text(text: str) -> str:
    if text is None:
        return ""
    return " ".join(text.split()).strip()


def get_text(tag) -> str:
    if tag is None:
        return ""
    return clean_text(tag.get_text(" ", strip=True))


def safe_name(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = text.replace("::", "_")
    text = re.sub(r"[^A-Za-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def normalize_fn_name(name: str) -> str:
    name = clean_text(name)
    name = re.sub(r"\([^)]*\)", "", name)

    if "::" in name:
        name = name.split("::")[-1]

    return name.strip().lower()


def extract_func_name(subprogram: str) -> str:
    subprogram = clean_text(subprogram)
    subprogram = re.sub(r"\([^)]*\)", "", subprogram)

    if "::" in subprogram:
        return subprogram.split("::")[-1].strip()

    return subprogram.strip()


def read_html_soup(html_path: Path):
    try:
        html = html_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        html = html_path.read_text(encoding="cp949", errors="ignore")

    return BeautifulSoup(html, "html.parser")


# ================================================================
# HTML MODIFIER - DATE STATE
# ================================================================

class DateState:
    def __init__(self):
        self.creation_dt = parse_report_date(TEST_START_DATE)
        self.execution_dt = parse_report_date(EXECUTION_START_DATE)
        self.tc_count = 0

    def next_creation_date(self) -> str:
        current = self.creation_dt
        step = random.randint(CREATION_MIN_STEP_MINUTES, CREATION_MAX_STEP_MINUTES)
        self.creation_dt += timedelta(minutes=step)
        return format_report_date(current)

    def next_execution_date(self) -> str:
        current = self.execution_dt

        self.tc_count += 1
        if self.tc_count % EXECUTION_STEP_PER_TC_COUNT == 0:
            self.execution_dt += timedelta(minutes=EXECUTION_STEP_MINUTES)

        return format_report_date(current)


def parse_report_date(text: str) -> datetime.datetime:
    return datetime.datetime.strptime(text, DATE_FORMAT)


def format_report_date(dt: datetime.datetime) -> str:
    return dt.strftime(DATE_FORMAT).upper()


# ================================================================
# HTML MODIFIER - COMPONENT / UNIT CANDIDATE LOGIC
# ================================================================

def build_component_unit_candidates(
    file_name: str,
    layer: str,
    parsed_component: str = "",
    parsed_unit: str = "",
) -> List[Tuple[str, str]]:
    candidates: List[Tuple[str, str]] = []

    if parsed_component and parsed_unit:
        candidates.append((parsed_component.strip(), parsed_unit.strip()))

    if not file_name or not layer:
        return candidates

    middle_match = re.match(
        rf"^{re.escape(layer)}_(?P<middle>.+)_UT_Report\.html$",
        file_name,
        re.IGNORECASE,
    )

    if not middle_match:
        return candidates

    middle = middle_match.group("middle").strip()
    parts = [p for p in middle.split("_") if p]

    for split_idx in range(1, len(parts)):
        comp = "_".join(parts[:split_idx])
        unit = "_".join(parts[split_idx:])
        pair = (comp, unit)

        if pair not in candidates:
            candidates.append(pair)

    return candidates


def find_required_functions_by_candidates(
    layer_map: Dict[str, Dict[str, List[str]]],
    candidates: List[Tuple[str, str]],
) -> Tuple[List[str], Optional[Tuple[str, str]]]:
    for component, unit in candidates:
        comp_key = component.strip().lower()
        unit_key = unit.strip().lower()

        required = layer_map.get(comp_key, {}).get(unit_key, [])
        if required:
            return required, (component, unit)

    return [], None


def find_uncovered_required_functions(
    required_functions: List[str],
    covered_functions: Set[str],
) -> List[str]:
    uncovered: List[str] = []

    for fn in required_functions:
        key = normalize_fn_name(fn)
        full_key = clean_text(fn).lower()

        if key and key not in covered_functions and full_key not in covered_functions:
            uncovered.append(fn)

    return uncovered


# ================================================================
# HTML MODIFIER - JSON LOGIC
# ================================================================

def get_layer_from_file_name(file_name: str) -> str:
    if "_" not in file_name:
        return ""

    return file_name.split("_", 1)[0].strip()


def get_json_path_by_html_name(html_file_name: str) -> Optional[Path]:
    name = html_file_name.strip().lower()

    for layer_key, json_name in JSON_FILE_MAP.items():
        if name.startswith(layer_key.lower() + "_"):
            return Path(SWE4_JSON_ROOT) / json_name

    return None


def load_function_list_json(path: Optional[Path]) -> Dict[str, Dict[str, List[str]]]:
    """
    Expected JSON:
    {
      "ComponentName": {
        "UnitName": [
          "FunctionA",
          "FunctionB"
        ]
      }
    }
    """
    if path is None:
        return {}

    if not path.exists():
        print(f"[WARN] JSON file not found: {path}")
        return {}

    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as e:
        print(f"[WARN] Failed to load JSON: {path} / {e}")
        return {}

    layer_map: Dict[str, Dict[str, List[str]]] = {}

    for comp, unit_map in raw.items():
        if not isinstance(unit_map, dict):
            continue

        comp_key = str(comp).strip().lower()
        if not comp_key:
            continue

        layer_map.setdefault(comp_key, {})

        for unit, funcs in unit_map.items():
            if not isinstance(funcs, list):
                continue

            unit_key = str(unit).strip().lower()
            if not unit_key:
                continue

            layer_map[comp_key][unit_key] = [
                str(fn).strip()
                for fn in funcs
                if str(fn).strip()
            ]

    return layer_map


def load_function_list_json_cached(path: Optional[Path]) -> Dict[str, Dict[str, List[str]]]:
    if path is None:
        return {}

    key = str(path).lower()

    if key in JSON_CACHE:
        return JSON_CACHE[key]

    data = load_function_list_json(path)
    JSON_CACHE[key] = data
    return data


# ================================================================
# HTML MODIFIER - HTML HELPERS
# ================================================================

def find_parent_report_block(tag):
    cur = tag

    while cur is not None:
        if getattr(cur, "name", None) == "div":
            classes = cur.get("class", [])
            if "report-block" in classes:
                return cur

        cur = cur.parent

    return None


def find_parent_contents_block(tag):
    cur = tag

    while cur is not None:
        if getattr(cur, "name", None) == "div":
            classes = cur.get("class", [])
            if "contents-block" in classes:
                return cur

        cur = cur.parent

    return None


def get_table_value_by_header(table, header_name: str) -> str:
    if table is None:
        return ""

    for tr in table.find_all("tr"):
        th = tr.find("th")
        td = tr.find("td")

        if not th or not td:
            continue

        if clean_text(th.get_text()) == header_name:
            return clean_text(td.get_text())

    return ""


def set_table_value_by_header(table, header_name: str, value: str) -> None:
    if table is None:
        return

    for tr in table.find_all("tr"):
        th = tr.find("th")
        td = tr.find("td")

        if not th or not td:
            continue

        if clean_text(th.get_text()) == header_name:
            td.clear()
            td.append(value)
            return


def find_config_blocks(soup: BeautifulSoup):
    """
    Test Case Configuration blocks.
    실제 HTML 등장 순서를 유지한다.
    """
    blocks = []

    for a in soup.find_all("a", id=re.compile(r"^TestCaseConfiguration_\d+$")):
        h3 = a.find_parent("h3")
        block = find_parent_report_block(h3 or a)

        if not block:
            continue

        table = block.find("table")
        if not table:
            continue

        m = re.match(r"^TestCaseConfiguration_(\d+)$", a.get("id", ""))
        if not m:
            continue

        idx = m.group(1)
        blocks.append((idx, block, table))

    return blocks


def collect_covered_functions(soup: BeautifulSoup) -> Set[str]:
    """
    Test Case Configuration의 Subprogram 기준으로 covered function set 생성.
    전체 td scan은 하지 않는다.
    """
    covered: Set[str] = set()

    for _, _, table in find_config_blocks(soup):
        subprogram = get_table_value_by_header(table, "Subprogram")

        if subprogram:
            covered.add(clean_text(subprogram).lower())
            covered.add(normalize_fn_name(subprogram))

    return covered


# ================================================================
# HTML MODIFIER - USER CODE SECTION
# ================================================================

def replace_user_code_section(
    soup: BeautifulSoup,
    uncovered: List[str],
    matched_pair: Optional[Tuple[str, str]],
) -> None:
    user_anchor = soup.find("a", id="UserCode")
    if not user_anchor:
        return

    block = find_parent_report_block(user_anchor)
    if not block:
        return

    row = block.find("div", class_="row")

    block.clear()

    if row:
        block.append(row)

    p = soup.new_tag("p")
    p.string = "Inline Function Coverage Inspection"
    block.append(p)

    if matched_pair:
        comp, unit = matched_pair

        info = soup.new_tag("p")
        info.string = f"Component/Unit mapping: {comp} / {unit}"
        block.append(info)

    if uncovered:
        ol = soup.new_tag("ol")

        for fn in uncovered:
            li = soup.new_tag("li")
            li.string = f'Inline Function "{fn}" has been covered by inspection.'
            ol.append(li)

        block.append(ol)
    else:
        p2 = soup.new_tag("p")
        p2.string = "No uncovered required inline function was found from the interface mapping."
        block.append(p2)


# ================================================================
# HTML MODIFIER - TC DISPLAY NAME / DATE / TOC
# ================================================================

def make_tc_display_name(
    uut: str,
    subprogram: str,
    counter: Dict[str, int],
) -> str:
    """
    Display name only.

    Example:
      HERMES::BootloaderManagement::BootloaderManagement
      -> BootloaderManagement-1
    """
    fn = safe_name(extract_func_name(subprogram))

    if not fn:
        fn = "Function"

    key = fn.lower()
    counter[key] = counter.get(key, 0) + 1

    return f"{fn}-{counter[key]}"


def get_toc_testcases_ul(soup: BeautifulSoup):
    """
    TOC 안의 Test Cases 하위 ul을 찾는다.

    구조:
      div.contents-block
        ul.toc-level1
          li.collapsible-toc title="Test Cases"
            ul
              li.tc-item ...
    """
    toc_anchor = soup.find("a", id="TableOfContents")
    if not toc_anchor:
        return None

    toc_block = find_parent_contents_block(toc_anchor)
    if not toc_block:
        return None

    testcases_li = toc_block.find(
        "li",
        class_=lambda c: c and "collapsible-toc" in c.split(),
        title="Test Cases",
    )

    if not testcases_li:
        return None

    return testcases_li.find("ul", recursive=False)


def update_table_of_contents_name(
    soup: BeautifulSoup,
    idx: str,
    new_name: str,
) -> None:
    tc_ul = get_toc_testcases_ul(soup)

    if tc_ul:
        a = tc_ul.find("a", href=re.compile(rf"^#TestCase_{re.escape(idx)}_"))
        if a:
            a.clear()
            a.append(new_name)

            li = a.find_parent("li", class_=lambda c: c and "tc-item" in c.split())
            if li and li.has_attr("title"):
                li["title"] = new_name
            return

    # fallback
    for a in soup.find_all("a", href=True):
        href = a.get("href", "")

        if href.startswith(f"#TestCase_{idx}_"):
            a.clear()
            a.append(new_name)

            li = a.find_parent("li")
            if li and li.has_attr("title"):
                li["title"] = new_name


def update_testcase_header_name(
    soup: BeautifulSoup,
    idx: str,
    new_name: str,
) -> None:
    testcase_anchor = soup.find("a", id=re.compile(rf"^TestCase_{idx}_"))
    if not testcase_anchor:
        return

    testcase_div = testcase_anchor.find_parent("div", class_="testcase")
    if not testcase_div:
        return

    span = testcase_div.find("span", class_="testcase_name")
    if span:
        span.clear()
        span.append(new_name)


def get_toc_li_by_idx(soup: BeautifulSoup, idx: str):
    """
    TestCase index 기준으로 TOC li를 찾는다.

    Example:
      idx = "133"
      href = "#TestCase_133_ATG-TEST-1"
    """
    tc_ul = get_toc_testcases_ul(soup)
    if not tc_ul:
        return None

    a = tc_ul.find("a", href=re.compile(rf"^#TestCase_{re.escape(idx)}_"))
    if not a:
        return None

    return a.find_parent("li", class_=lambda c: c and "tc-item" in c.split())


def reorder_toc_testcase_entries(
    soup: BeautifulSoup,
    ordered_tc_indices: List[str],
) -> None:
    """
    TOC의 Test Cases 항목을 실제 Test Case Configuration 처리 순서대로 재배치한다.
    """
    if not ordered_tc_indices:
        return

    tc_ul = get_toc_testcases_ul(soup)
    if not tc_ul:
        print("[WARN] TOC Test Cases ul not found")
        return

    ordered_lis = []

    for idx in ordered_tc_indices:
        li = get_toc_li_by_idx(soup, idx)
        if li:
            ordered_lis.append(li)
        else:
            print(f"[WARN] TOC li not found for TestCase index: {idx}")

    if not ordered_lis:
        print("[WARN] No TOC li items were collected")
        return

    existing_tc_lis = [
        li for li in tc_ul.find_all("li", recursive=False)
        if "tc-item" in li.get("class", [])
    ]

    for li in existing_tc_lis:
        li.extract()

    for li in ordered_lis:
        tc_ul.append(li)


def update_tc_display_names_and_dates(
    soup: BeautifulSoup,
    date_state: DateState,
) -> None:
    tc_counter: Dict[str, int] = {}
    ordered_tc_indices: List[str] = []

    for idx, _, table in find_config_blocks(soup):
        ordered_tc_indices.append(idx)

        uut = get_table_value_by_header(table, "Unit Under Test")
        subprogram = get_table_value_by_header(table, "Subprogram")
        old_name = get_table_value_by_header(table, "Test Case Name")

        if not uut:
            uut = "UT"

        if not subprogram:
            subprogram = old_name or "Function"

        display_name = make_tc_display_name(uut, subprogram, tc_counter)

        # Keep original Test Case Name in Test Case Configuration.
        set_table_value_by_header(table, "Date of Creation", date_state.next_creation_date())
        set_table_value_by_header(table, "Date of Execution", date_state.next_execution_date())

        # Change display title and TOC only.
        update_table_of_contents_name(soup, idx, display_name)
        update_testcase_header_name(soup, idx, display_name)

    reorder_toc_testcase_entries(soup, ordered_tc_indices)


# ================================================================
# HTML MODIFIER - REQUIREMENTS / NOTES CLEANUP
# ================================================================

def clean_requirements_notes_text(text: str) -> str:
    lines = text.splitlines()
    new_lines: List[str] = []

    skip_next_atg_line = False

    for line in lines:
        stripped = line.strip()

        if stripped == "This is an automatically generated test case.":
            continue

        if stripped == "Test Case Generation Notes:":
            skip_next_atg_line = True
            continue

        if stripped in {
            "ATG constructed a complete test-case.",
            "ATG constructed a partial test-case.",
        }:
            continue

        if skip_next_atg_line and stripped.startswith("ATG constructed"):
            skip_next_atg_line = False
            continue

        skip_next_atg_line = False
        new_lines.append(line.rstrip())

    compact: List[str] = []
    blank_count = 0

    for line in new_lines:
        if line.strip() == "":
            blank_count += 1
            if blank_count > 1:
                continue
        else:
            blank_count = 0

        compact.append(line)

    return "\n".join(compact).strip()


def clean_requirements_notes(soup: BeautifulSoup) -> None:
    for h4 in soup.find_all("h4"):
        if clean_text(h4.get_text()) != "Requirements/Notes":
            continue

        pre = h4.find_next("pre")
        if not pre:
            continue

        cleaned = clean_requirements_notes_text(pre.get_text())
        pre.clear()
        pre.append(cleaned)


# ================================================================
# HTML MODIFIER - PROCESS ONE HTML
# ================================================================

def process_html_file(
    html_path: Path,
    date_state: DateState,
) -> None:
    total_t0 = time.perf_counter()

    t0 = time.perf_counter()
    soup = read_html_soup(html_path)

    if ENABLE_TIME_LOG:
        print(f"  read+parse: {time.perf_counter() - t0:.2f}s")

    file_name = html_path.name
    layer = get_layer_from_file_name(file_name)

    t0 = time.perf_counter()
    json_path = get_json_path_by_html_name(file_name)
    layer_map = load_function_list_json_cached(json_path)

    if ENABLE_TIME_LOG:
        print(f"  json: {time.perf_counter() - t0:.2f}s")

    candidates = build_component_unit_candidates(
        file_name=file_name,
        layer=layer,
    )

    t0 = time.perf_counter()
    covered_functions = collect_covered_functions(soup)

    if ENABLE_TIME_LOG:
        print(f"  collect covered: {time.perf_counter() - t0:.2f}s")

    required_functions, matched_pair = find_required_functions_by_candidates(
        layer_map=layer_map,
        candidates=candidates,
    )

    uncovered = find_uncovered_required_functions(
        required_functions=required_functions,
        covered_functions=covered_functions,
    )

    t0 = time.perf_counter()

    replace_user_code_section(
        soup=soup,
        uncovered=uncovered,
        matched_pair=matched_pair,
    )

    update_tc_display_names_and_dates(
        soup=soup,
        date_state=date_state,
    )

    clean_requirements_notes(soup)

    if ENABLE_TIME_LOG:
        print(f"  modify: {time.perf_counter() - t0:.2f}s")

    t0 = time.perf_counter()
    html_path.write_text(str(soup), encoding="utf-8")

    if ENABLE_TIME_LOG:
        print(f"  write: {time.perf_counter() - t0:.2f}s")
        print(f"  total: {time.perf_counter() - total_t0:.2f}s")


# ================================================================
# HTML MODIFIER - COPY AND PROCESS
# ================================================================

def copy_original_to_new() -> None:
    src = Path(ORIGINAL_ROOT)
    dst = Path(NEW_ROOT)

    if not src.exists():
        raise FileNotFoundError(f"Original root does not exist: {src}")

    if dst.exists():
        if OVERWRITE_NEW_ROOT:
            shutil.rmtree(dst)
        else:
            raise FileExistsError(f"New root already exists: {dst}")

    shutil.copytree(src, dst)


def collect_html_files(input_root: str) -> List[Path]:
    root = Path(input_root)

    if not root.exists():
        raise FileNotFoundError(f"Input root does not exist: {root}")

    if not root.is_dir():
        raise NotADirectoryError(f"Input root is not a folder: {root}")

    if RECURSIVE_SEARCH:
        files = [
            p for p in root.rglob("*")
            if p.is_file() and p.suffix.lower() in [".html", ".htm"]
        ]
    else:
        files = [
            p for p in root.iterdir()
            if p.is_file() and p.suffix.lower() in [".html", ".htm"]
        ]

    files.sort(key=lambda p: str(p).lower())
    return files


def process_all_html_files() -> None:
    random.seed(RANDOM_SEED)

    print("[COPY] Original folder to new folder")
    copy_t0 = time.perf_counter()

    copy_original_to_new()

    print(f"[COPY DONE] {time.perf_counter() - copy_t0:.2f}s")

    date_state = DateState()

    html_files = collect_html_files(NEW_ROOT)

    for html_path in html_files:
        print(f"[PROCESS] {html_path}")
        process_html_file(
            html_path=html_path,
            date_state=date_state,
        )

    print(f"[DONE] Processed HTML files: {len(html_files)}")
    print(f"[OUTPUT] {NEW_ROOT}")


# ================================================================
# EXCEL EXTRACTOR - CREATE / STYLE
# ================================================================

def create_excel(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TCNames"

    ws.cell(row=1, column=1, value="Source HTML")
    ws.cell(row=1, column=2, value="TC-Count")
    ws.cell(row=1, column=3, value="Unit Under Test")
    ws.cell(row=1, column=4, value="Function Name")
    ws.cell(row=1, column=5, value="TestCase Name")

    ws.cell(row=1, column=6, value="Input")
    ws.cell(row=2, column=6, value="Variable")
    ws.cell(row=2, column=7, value="DataType")
    ws.cell(row=2, column=8, value="Value")

    ws.cell(row=1, column=9, value="Output")
    ws.cell(row=2, column=9, value="Variable")
    ws.cell(row=2, column=10, value="DataType")
    ws.cell(row=2, column=11, value="Value")

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:H1")
    ws.merge_cells("I1:K1")

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=11):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"

    wb.save(output_path)


def auto_fit_columns(ws):
    max_width_map = {
        1: 45,
        2: 12,
        3: 35,
        4: 45,
        5: 35,
        6: 45,
        7: 35,
        8: 50,
        9: 45,
        10: 35,
        11: 50,
    }

    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            if cell.value is None:
                continue

            value = str(cell.value)
            max_length = max(max_length, len(value))

        width = min(max_length + 2, max_width_map.get(col_idx, 40))
        ws.column_dimensions[col_letter].width = width


def style_body_range(ws):
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


# ================================================================
# EXCEL EXTRACTOR - HTML PARSING
# ================================================================

def parse_summary_table(testcase_div):
    unit_under_test = ""
    subprogram = ""

    first_table = testcase_div.find("table")
    if not first_table:
        return unit_under_test, subprogram

    rows = first_table.find_all("tr")

    if len(rows) >= 1:
        tds = rows[0].find_all("td")
        if tds:
            unit_under_test = get_text(tds[0])

    if len(rows) >= 2:
        tds = rows[1].find_all("td")
        if tds:
            subprogram = get_text(tds[0])

    return unit_under_test, subprogram


def extract_table_rows(table):
    rows_out = []
    tbody = table.find("tbody") or table
    rows = tbody.find_all("tr", recursive=False)

    for row in rows:
        tds = row.find_all("td", recursive=False)
        vals = [get_text(td) for td in tds]

        while len(vals) < 3:
            vals.append("")

        rows_out.append((vals[0], vals[1], vals[2]))

    return rows_out


def filter_meaningful_rows(raw_rows, section_name=""):
    filtered = []

    kept_instance_header = False
    keep_next_type_after_instance = False
    keep_next_type_after_return = False

    for c1, c2, c3 in raw_rows:
        text = c1.strip()
        has_data = c2.strip() != "" or c3.strip() != ""

        if text == "" and not has_data:
            continue

        if has_data:
            if text.startswith("UUT:") or text.startswith("Unit:") or text.startswith("Globals:"):
                continue
            filtered.append((c1, c2, c3))
            continue

        if text == "class members":
            continue

        if text.startswith("UUT:") or text.startswith("Unit:") or text.startswith("Globals:"):
            continue

        if text.startswith("Subprogram:"):
            filtered.append((c1, c2, c3))
            continue

        if text.startswith("Stubbed Subprograms"):
            filtered.append((c1, c2, c3))
            continue

        if not kept_instance_header and text.endswith(" Instance"):
            filtered.append((c1, c2, c3))
            kept_instance_header = True
            keep_next_type_after_instance = True
            continue

        if keep_next_type_after_instance:
            filtered.append((c1, c2, c3))
            keep_next_type_after_instance = False
            continue

        if text == "return":
            filtered.append((c1, c2, c3))
            keep_next_type_after_return = True
            continue

        if keep_next_type_after_return:
            filtered.append((c1, c2, c3))
            keep_next_type_after_return = False
            continue

    return filtered


def compress_return_rows(rows):
    compressed = []
    i = 0

    while i < len(rows):
        c1, c2, c3 = rows[i]
        text = c1.strip()

        if text == "return" and c2.strip() == "" and c3.strip() == "":
            if i + 1 < len(rows):
                n1, n2, n3 = rows[i + 1]
                next_text = n1.strip()

                if (
                    next_text != ""
                    and n2.strip() == ""
                    and n3.strip() == ""
                    and not next_text.startswith("Subprogram:")
                    and not next_text.startswith("Stubbed Subprograms")
                    and not next_text.endswith(" Instance")
                    and next_text != "return"
                ):
                    compressed.append(("return", next_text, ""))
                    i += 2
                    continue

        compressed.append(rows[i])
        i += 1

    return compressed


def parse_test_data_section(testcase_div):
    result = {
        "Input Test Data": [],
        "Expected Test Data": []
    }

    direct_children = testcase_div.find_all("div", recursive=False)

    for block in direct_children:
        h3 = block.find("h3", recursive=False)
        if not h3:
            continue

        if get_text(h3) != "Test Case Data":
            continue

        subsections = block.find_all("div", recursive=False)

        for subsection in subsections:
            h4 = subsection.find("h4", recursive=False)
            if not h4:
                continue

            section_name = get_text(h4)
            if section_name not in result:
                continue

            table = subsection.find("table")
            if not table:
                continue

            raw_rows = extract_table_rows(table)
            filtered_rows = filter_meaningful_rows(raw_rows, section_name)
            compressed_rows = compress_return_rows(filtered_rows)

            result[section_name] = compressed_rows

        break

    return result


def find_main_scroller(soup):
    return soup.find(id="main-scroller")


def find_testcases(main_scroller):
    if not main_scroller:
        return []
    return main_scroller.find_all("div", class_="testcase", recursive=False)


# ================================================================
# EXCEL EXTRACTOR - WRITE
# ================================================================

def write_testcase_header(
    ws,
    row_idx,
    source_html,
    tc_count,
    unit_under_test,
    subprogram,
    tc_name
):
    ws.cell(row=row_idx, column=1, value=source_html)
    ws.cell(row=row_idx, column=2, value=tc_count)
    ws.cell(row=row_idx, column=3, value=unit_under_test)
    ws.cell(row=row_idx, column=4, value=subprogram)
    ws.cell(row=row_idx, column=5, value=tc_name)


def write_data_block(ws, start_row, start_col, items):
    row = start_row

    for var_name, data_type, value in items:
        ws.cell(row=row, column=start_col, value=var_name)
        ws.cell(row=row, column=start_col + 1, value=data_type)
        ws.cell(row=row, column=start_col + 2, value=value)
        row += 1

    return row


def append_testcase_to_sheet(
    ws,
    source_html,
    tc_count,
    tc_name,
    unit_under_test,
    subprogram,
    input_items,
    expected_items
):
    start_row = ws.max_row + 2

    write_testcase_header(
        ws=ws,
        row_idx=start_row,
        source_html=source_html,
        tc_count=tc_count,
        unit_under_test=unit_under_test,
        subprogram=subprogram,
        tc_name=tc_name
    )

    input_end = write_data_block(ws, start_row, 6, input_items)
    expected_end = write_data_block(ws, start_row, 9, expected_items)

    return max(input_end, expected_end)


# ================================================================
# EXCEL EXTRACTOR - BULK PROCESS
# ================================================================

def parse_html_file_to_sheet(html_path: Path, ws, global_tc_count: int) -> int:
    soup = read_html_soup(html_path)

    main_scroller = find_main_scroller(soup)
    if main_scroller is None:
        logging.warning("main-scroller not found: %s", html_path)
        print(f"[WARN] main-scroller not found: {html_path}")
        return global_tc_count

    testcases = find_testcases(main_scroller)

    logging.info("Found %d testcases in %s", len(testcases), html_path)
    print(f"  testcases: {len(testcases)}")

    for testcase_div in testcases:
        try:
            global_tc_count += 1

            h2 = testcase_div.find("h2")
            tc_name = ""
            if h2:
                span = h2.find("span")
                tc_name = get_text(span if span else h2)

            unit_under_test, subprogram = parse_summary_table(testcase_div)

            test_data = parse_test_data_section(testcase_div)
            input_items = test_data.get("Input Test Data", [])
            expected_items = test_data.get("Expected Test Data", [])

            append_testcase_to_sheet(
                ws=ws,
                source_html=html_path.name,
                tc_count=global_tc_count,
                tc_name=tc_name,
                unit_under_test=unit_under_test,
                subprogram=subprogram,
                input_items=input_items,
                expected_items=expected_items
            )

        except Exception:
            logging.exception("Failed while parsing testcase in file: %s", html_path)

    return global_tc_count


def process_folder_to_excel():
    start_time = datetime.datetime.now()
    total_t0 = time.perf_counter()

    print(f"[START] {start_time}")
    print(f"[INPUT] {INPUT_ROOT}")
    print(f"[OUTPUT] {OUTPUT_XLSX_PATH}")

    output_path = Path(OUTPUT_XLSX_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    create_excel(str(output_path))

    wb = openpyxl.load_workbook(output_path)
    ws = wb["TCNames"]

    html_files = collect_html_files(INPUT_ROOT)

    if not html_files:
        print("[WARN] No HTML files found.")
        wb.save(output_path)
        return

    print(f"[FOUND] HTML files: {len(html_files)}")

    global_tc_count = 0

    for html_path in html_files:
        file_t0 = time.perf_counter()
        print(f"[PROCESS] {html_path}")

        try:
            global_tc_count = parse_html_file_to_sheet(
                html_path=html_path,
                ws=ws,
                global_tc_count=global_tc_count
            )

            if ENABLE_TIME_LOG:
                print(f"  elapsed: {time.perf_counter() - file_t0:.2f}s")

        except Exception:
            logging.exception("Failed while processing file: %s", html_path)
            print(f"[ERROR] Failed while processing: {html_path}")

    style_body_range(ws)
    auto_fit_columns(ws)

    wb.save(output_path)

    end_time = datetime.datetime.now()

    print(f"[DONE] HTML files processed: {len(html_files)}")
    print(f"[DONE] Testcases extracted: {global_tc_count}")
    print(f"[DONE] Output: {output_path}")
    print(f"[END] {end_time}")
    print(f"[TOTAL] {time.perf_counter() - total_t0:.2f}s")


# ================================================================
# MAIN
# ================================================================

def main():
    random.seed(RANDOM_SEED)

    # ============================================================
    # OPTION 1. Run both steps
    # ============================================================
    process_all_html_files()
    process_folder_to_excel()

    # ============================================================
    # OPTION 2. Run HTML modification only
    # ============================================================
    # process_all_html_files()

    # ============================================================
    # OPTION 3. Run Excel extraction only
    # ============================================================
    # process_folder_to_excel()


if __name__ == "__main__":
    main()