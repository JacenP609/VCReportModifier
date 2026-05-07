import datetime
import logging
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter

from report_modifier import clean_text, get_text, collect_html_files, read_html_soup

INPUT_ROOT=""
OUTPUT_XLSX_PATH=""
ENABLE_TIME_LOG=True
RECURSIVE_SEARCH=True

HEADER_FILL = PatternFill("solid", start_color="38E3FF")
HEADER_FONT = Font(bold=True, size=12, color="0000FF")
THIN_SIDE = Side(border_style="thin", color="000000")
HEADER_BORDER = Border(
    top=THIN_SIDE,
    left=THIN_SIDE,
    right=THIN_SIDE,
    bottom=THIN_SIDE
)

BODY_BORDER = Border(
    top=THIN_SIDE,
    left=THIN_SIDE,
    right=THIN_SIDE,
    bottom=THIN_SIDE
)

# ================================================================
# EXCEL EXTRACTOR - CREATE / STYLE
# ================================================================

def create_excel(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TCNames"

    ws.cell(row=1, column=1, value="Source HTML")
    ws.cell(row=1, column=2, value="TC-Count")
    ws.cell(row=1, column=3, value="Unit Under Test")
    ws.cell(row=1, column=4, value="Function Name")
    ws.cell(row=1, column=5, value="TestCase Name")

    ws.cell(row=1, column=6, value="Input")
    ws.cell(row=2, column=6, value="Variable")
    ws.cell(row=2, column=7, value="DataType")
    ws.cell(row=2, column=8, value="Value")

    ws.cell(row=1, column=9, value="Output")
    ws.cell(row=2, column=9, value="Variable")
    ws.cell(row=2, column=10, value="DataType")
    ws.cell(row=2, column=11, value="Value")

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:H1")
    ws.merge_cells("I1:K1")

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=11):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"

    wb.save(output_path)


def auto_fit_columns(ws):
    max_width_map = {
        1: 45,
        2: 12,
        3: 35,
        4: 45,
        5: 35,
        6: 45,
        7: 35,
        8: 50,
        9: 45,
        10: 35,
        11: 50,
    }

    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            if cell.value is None:
                continue

            value = str(cell.value)
            max_length = max(max_length, len(value))

        width = min(max_length + 2, max_width_map.get(col_idx, 40))
        ws.column_dimensions[col_letter].width = width


def style_body_range(ws):
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


# ================================================================
# EXCEL EXTRACTOR - HTML PARSING
# ================================================================

def parse_summary_table(testcase_div):
    unit_under_test = ""
    subprogram = ""

    first_table = testcase_div.find("table")
    if not first_table:
        return unit_under_test, subprogram

    rows = first_table.find_all("tr")

    if len(rows) >= 1:
        tds = rows[0].find_all("td")
        if tds:
            unit_under_test = get_text(tds[0])

    if len(rows) >= 2:
        tds = rows[1].find_all("td")
        if tds:
            subprogram = get_text(tds[0])

    return unit_under_test, subprogram


def extract_table_rows(table):
    rows_out = []
    tbody = table.find("tbody") or table
    rows = tbody.find_all("tr", recursive=False)

    for row in rows:
        tds = row.find_all("td", recursive=False)
        vals = [get_text(td) for td in tds]

        while len(vals) < 3:
            vals.append("")

        rows_out.append((vals[0], vals[1], vals[2]))

    return rows_out


def filter_meaningful_rows(raw_rows, section_name=""):
    filtered = []

    kept_instance_header = False
    keep_next_type_after_instance = False
    keep_next_type_after_return = False

    for c1, c2, c3 in raw_rows:
        text = c1.strip()
        has_data = c2.strip() != "" or c3.strip() != ""

        if text == "" and not has_data:
            continue

        if has_data:
            if text.startswith("UUT:") or text.startswith("Unit:") or text.startswith("Globals:"):
                continue
            filtered.append((c1, c2, c3))
            continue

        if text == "class members":
            continue

        if text.startswith("UUT:") or text.startswith("Unit:") or text.startswith("Globals:"):
            continue

        if text.startswith("Subprogram:"):
            filtered.append((c1, c2, c3))
            continue

        if text.startswith("Stubbed Subprograms"):
            filtered.append((c1, c2, c3))
            continue

        if not kept_instance_header and text.endswith(" Instance"):
            filtered.append((c1, c2, c3))
            kept_instance_header = True
            keep_next_type_after_instance = True
            continue

        if keep_next_type_after_instance:
            filtered.append((c1, c2, c3))
            keep_next_type_after_instance = False
            continue

        if text == "return":
            filtered.append((c1, c2, c3))
            keep_next_type_after_return = True
            continue

        if keep_next_type_after_return:
            filtered.append((c1, c2, c3))
            keep_next_type_after_return = False
            continue

    return filtered


def compress_return_rows(rows):
    compressed = []
    i = 0

    while i < len(rows):
        c1, c2, c3 = rows[i]
        text = c1.strip()

        if text == "return" and c2.strip() == "" and c3.strip() == "":
            if i + 1 < len(rows):
                n1, n2, n3 = rows[i + 1]
                next_text = n1.strip()

                if (
                    next_text != ""
                    and n2.strip() == ""
                    and n3.strip() == ""
                    and not next_text.startswith("Subprogram:")
                    and not next_text.startswith("Stubbed Subprograms")
                    and not next_text.endswith(" Instance")
                    and next_text != "return"
                ):
                    compressed.append(("return", next_text, ""))
                    i += 2
                    continue

        compressed.append(rows[i])
        i += 1

    return compressed


def parse_test_data_section(testcase_div):
    result = {
        "Input Test Data": [],
        "Expected Test Data": []
    }

    direct_children = testcase_div.find_all("div", recursive=False)

    for block in direct_children:
        h3 = block.find("h3", recursive=False)
        if not h3:
            continue

        if get_text(h3) != "Test Case Data":
            continue

        subsections = block.find_all("div", recursive=False)

        for subsection in subsections:
            h4 = subsection.find("h4", recursive=False)
            if not h4:
                continue

            section_name = get_text(h4)
            if section_name not in result:
                continue

            table = subsection.find("table")
            if not table:
                continue

            raw_rows = extract_table_rows(table)
            filtered_rows = filter_meaningful_rows(raw_rows, section_name)
            compressed_rows = compress_return_rows(filtered_rows)

            result[section_name] = compressed_rows

        break

    return result


def find_main_scroller(soup):
    return soup.find(id="main-scroller")


def find_testcases(main_scroller):
    if not main_scroller:
        return []
    return main_scroller.find_all("div", class_="testcase", recursive=False)


# ================================================================
# EXCEL EXTRACTOR - WRITE
# ================================================================

def write_testcase_header(
    ws,
    row_idx,
    source_html,
    tc_count,
    unit_under_test,
    subprogram,
    tc_name
):
    ws.cell(row=row_idx, column=1, value=source_html)
    ws.cell(row=row_idx, column=2, value=tc_count)
    ws.cell(row=row_idx, column=3, value=unit_under_test)
    ws.cell(row=row_idx, column=4, value=subprogram)
    ws.cell(row=row_idx, column=5, value=tc_name)


def write_data_block(ws, start_row, start_col, items):
    row = start_row

    for var_name, data_type, value in items:
        ws.cell(row=row, column=start_col, value=var_name)
        ws.cell(row=row, column=start_col + 1, value=data_type)
        ws.cell(row=row, column=start_col + 2, value=value)
        row += 1

    return row


def append_testcase_to_sheet(
    ws,
    source_html,
    tc_count,
    tc_name,
    unit_under_test,
    subprogram,
    input_items,
    expected_items
):
    start_row = ws.max_row + 2

    write_testcase_header(
        ws=ws,
        row_idx=start_row,
        source_html=source_html,
        tc_count=tc_count,
        unit_under_test=unit_under_test,
        subprogram=subprogram,
        tc_name=tc_name
    )

    input_end = write_data_block(ws, start_row, 6, input_items)
    expected_end = write_data_block(ws, start_row, 9, expected_items)

    return max(input_end, expected_end)


# ================================================================
# EXCEL EXTRACTOR - BULK PROCESS
# ================================================================

def parse_html_file_to_sheet(html_path: Path, ws, global_tc_count: int) -> int:
    soup = read_html_soup(html_path)

    main_scroller = find_main_scroller(soup)
    if main_scroller is None:
        logging.warning("main-scroller not found: %s", html_path)
        print(f"[WARN] main-scroller not found: {html_path}")
        return global_tc_count

    testcases = find_testcases(main_scroller)

    logging.info("Found %d testcases in %s", len(testcases), html_path)
    print(f"  testcases: {len(testcases)}")

    for testcase_div in testcases:
        try:
            global_tc_count += 1

            h2 = testcase_div.find("h2")
            tc_name = ""
            if h2:
                span = h2.find("span")
                tc_name = get_text(span if span else h2)

            unit_under_test, subprogram = parse_summary_table(testcase_div)

            test_data = parse_test_data_section(testcase_div)
            input_items = test_data.get("Input Test Data", [])
            expected_items = test_data.get("Expected Test Data", [])

            append_testcase_to_sheet(
                ws=ws,
                source_html=html_path.name,
                tc_count=global_tc_count,
                tc_name=tc_name,
                unit_under_test=unit_under_test,
                subprogram=subprogram,
                input_items=input_items,
                expected_items=expected_items
            )

        except Exception:
            logging.exception("Failed while parsing testcase in file: %s", html_path)

    return global_tc_count


def process_folder_to_excel(config):
    global INPUT_ROOT, OUTPUT_XLSX_PATH, ENABLE_TIME_LOG, RECURSIVE_SEARCH
    INPUT_ROOT = config["NEW_ROOT"]
    OUTPUT_XLSX_PATH = config["OUTPUT_XLSX_PATH"]
    ENABLE_TIME_LOG = config.get("ENABLE_TIME_LOG", True)
    RECURSIVE_SEARCH = config["RECURSIVE_SEARCH"]

    start_time = datetime.datetime.now()
    total_t0 = time.perf_counter()

    print(f"[START] {start_time}")
    print(f"[INPUT] {INPUT_ROOT}")
    print(f"[OUTPUT] {OUTPUT_XLSX_PATH}")

    output_path = Path(OUTPUT_XLSX_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    create_excel(str(output_path))

    wb = openpyxl.load_workbook(output_path)
    ws = wb["TCNames"]

    html_files = collect_html_files(INPUT_ROOT)

    if not html_files:
        print("[WARN] No HTML files found.")
        wb.save(output_path)
        return

    print(f"[FOUND] HTML files: {len(html_files)}")

    global_tc_count = 0

    for html_path in html_files:
        file_t0 = time.perf_counter()
        print(f"[PROCESS] {html_path}")

        try:
            global_tc_count = parse_html_file_to_sheet(
                html_path=html_path,
                ws=ws,
                global_tc_count=global_tc_count
            )

            if ENABLE_TIME_LOG:
                print(f"  elapsed: {time.perf_counter() - file_t0:.2f}s")

        except Exception:
            logging.exception("Failed while processing file: %s", html_path)
            print(f"[ERROR] Failed while processing: {html_path}")

    style_body_range(ws)
    auto_fit_columns(ws)

    wb.save(output_path)

    end_time = datetime.datetime.now()

    print(f"[DONE] HTML files processed: {len(html_files)}")
    print(f"[DONE] Testcases extracted: {global_tc_count}")
    print(f"[DONE] Output: {output_path}")
    print(f"[END] {end_time}")
    print(f"[TOTAL] {time.perf_counter() - total_t0:.2f}s")




if __name__ == "__main__":
    process_folder_to_excel()
